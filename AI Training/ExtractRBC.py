import win32com.client
import pandas as pd
import re
from datetime import datetime
import openpyxl

def extract_content_up_to_marker(body, marker):
    marker_position = body.find(marker)
    if marker_position != -1:
        return body[:marker_position].strip()
    else:
        return body.strip()


def remove_sheet_if_exists(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])



def parse_line(line):
    size_first_pattern = r"(\d+(\.\d+)?(mm|m|k))\s+([\w\s-]+?)\s+\((\w+)\)\s+(\d+\.\d+)\s+(bid|offered|offer)\s*(?:@|at|-)?\s*(\d*\.\d+)?"
    name_first_pattern = r"([\w\s-]+?)\s+\((\w+)\)\s+(\d*\.\d+)?\s*(bid|offered|offer)\s*(?:@|at|-)?\s*(\d+\.\d+)"
    dual_action_pattern = r"(\d+(\.\d+)?(mm|m|k))\s+([\w\s-]+?)\s+\((\w+)\)\s+(\d+\.\d+)\s+(bid)\s+/\s+(\d+\.\d+)\s+(offer)"

    default_dict = {"Name": "", "Size": "", "CUSIP": "", "Actions": "", "Price": "", "Error": line}

    entries = []

    # Size-First Format
    if re.match(size_first_pattern, line):
        for match in re.finditer(size_first_pattern, line):
            size, name, cusip, price, action, alt_price = match.groups()[0], match.groups()[3], match.groups()[4], match.groups()[5], match.groups()[6], match.groups()[7]
            entries.append({"Name": name.strip(), "Size": size, "CUSIP": cusip, "Actions": action, "Price": price if price else alt_price, "Error": ""})

    # Name-First Format
    elif re.match(name_first_pattern, line):
        for match in re.finditer(name_first_pattern, line):
            name, cusip, alt_price, action, price = match.groups()[0], match.groups()[1], match.groups()[2], match.groups()[3], match.groups()[4]
            entries.append({"Name": name.strip(), "Size": "", "CUSIP": cusip, "Actions": action, "Price": price if price else alt_price, "Error": ""})

    # Dual-Action Format
    elif re.match(dual_action_pattern, line):
        for match in re.finditer(dual_action_pattern, line):
            size, name, cusip, bid_price, offer_price = match.groups()[0], match.groups()[3], match.groups()[4], match.groups()[5], match.groups()[7]
            entries.append({"Name": name.strip(), "Size": size, "CUSIP": cusip, "Actions": "bid", "Price": bid_price, "Error": ""})
            entries.append({"Name": name.strip(), "Size": size, "CUSIP": cusip, "Actions": "offer", "Price": offer_price, "Error": ""})

    return entries if entries else [default_dict]

def write_df_to_excel(writer, df, sheet_name):
    if sheet_name in writer.book.sheetnames:
        idx = writer.book.sheetnames.index(sheet_name)
        writer.book.remove(writer.book.worksheets[idx])
    df.to_excel(writer, sheet_name=sheet_name, index=False)

# Create an Outlook application object and access emails
Outlook = win32com.client.Dispatch("Outlook.Application")
namespace = Outlook.GetNamespace("MAPI")
inbox = namespace.GetDefaultFolder(6)  # 6 refers to the inbox
bhCatBondFolder = inbox.Folders["BH Cat Bond"]

emails = []
sorted_emails = []

for message in bhCatBondFolder.Items:
    if message.UnRead:
        subject = message.Subject
        body = message.Body
        received_time = message.ReceivedTime
        formatted_time = received_time.strftime('%Y-%m-%d %H:%M:%S')
        extracted_body = extract_content_up_to_marker(body, "Craig Bonder")

        emails.append({
            "Timestamp": formatted_time,
            "Subject": subject,
            "Content": extracted_body
        })

        lines = extracted_body.split('\n')
        for line in lines:
            entries = parse_line(line.strip())
            sorted_emails.extend(entries)

        # message.UnRead = False  # Uncomment to mark as read
        # message.Save()  # Uncomment to save the state

emails_df = pd.DataFrame(emails)
sorted_df = pd.DataFrame(sorted_emails)

# File path for the Excel file
excel_file_path = "//ad-its.credit-agricole.fr/Amundi_Boston/Homedirs/buonomo/@Config/Desktop/Outlook Scanner/OrganizedBondEntries.xlsx"

# Open the workbook and remove sheets if they exist
workbook = openpyxl.load_workbook(excel_file_path)
remove_sheet_if_exists(workbook, 'Unread Emails')
remove_sheet_if_exists(workbook, 'Sorted')
workbook.save(excel_file_path)
workbook.close()

# Save to Excel with two sheets
with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
    emails_df.to_excel(writer, sheet_name='Unread Emails', index=False)
    sorted_df.to_excel(writer, sheet_name='Sorted', index=False)

print("Emails processed and saved to Excel file.")
