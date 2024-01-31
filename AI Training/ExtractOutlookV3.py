import win32com.client
import pandas as pd
import re
from datetime import datetime
import openpyxl

def extract_content_up_to_marker(body, marker):
    marker_position = body.find(marker)
    if marker_position != -1:
        return body[:marker_position].strip()
    else:
        return body.strip()

def remove_sheet_if_exists(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

def parse_line(line):
    patterns = {
        "size_name_cusip_offered_at_price": r"(\d+(\.\d+)?(mm|m|k))\s+([\w\s-]+?)\s+\((\w+)\)\s+offered\s+(?:@|at)\s+(\d*\.\d+)",
        "name_cusip_bid_at_price": r"([\w\s-]+?)\s+\((\w+)\)\s+bid\s+(?:@|at)\s+(\d+\.\d+)",
        "size_name_cusip_bid_offer": r"(\d+(\.\d+)?(mm|m|k))\s+([\w\s-]+?)\s+\((\w+)\)\s+(\d+\.\d+)\s+bid\s+/\s+(\d+\.\d+)\s+offer",
        "name_cusip_offered_at_price_no_size": r"([\w\s-]+?)\s+\((\w+)\)\s+offered\s+(?:@|at)\s+(\d+\.\d+)",
        "bid_price_for_name_cusip": r"(\d+\.\d+)\s+bid\s+for\s+([\w\s-]+?)\s+\((\w+)\)",
        "cusip_first_bid_at_price": r"(\w+)\s+([\w\s-]+?)\s+bid\s+(?:@|at)\s+(\d+\.\d+)"
    }

    default_dict = {"Name": "", "Size": "", "CUSIP": "", "Actions": "", "Price": "", "Sentence": line, "Function": "No Match", "Error": line}
    entries = []

    for key, pattern in patterns.items():
        if re.match(pattern, line):
            for match in re.finditer(pattern, line):
                groups = match.groups()
                entry = {"Sentence": line, "Function": key, "Error": ""}
                if key in ["size_name_cusip_offered_at_price", "size_name_cusip_bid_offer"]:
                    entry.update({"Name": groups[3].strip(), "Size": groups[0], "CUSIP": groups[4], "Actions": "bid/offer", "Price": groups[-1]})
                elif key in ["name_cusip_bid_at_price", "name_cusip_offered_at_price_no_size", "cusip_first_bid_at_price"]:
                    entry.update({"Name": groups[1].strip(), "CUSIP": groups[0], "Actions": "bid/offer", "Price": groups[2]})
                elif key == "bid_price_for_name_cusip":
                    entry.update({"Name": groups[1].strip(), "CUSIP": groups[2], "Actions": "bid", "Price": groups[0]})
                entries.append(entry)

    return entries if entries else [default_dict]



def write_df_to_excel(writer, df, sheet_name):
    if sheet_name in writer.book.sheetnames:
        idx = writer.book.sheetnames.index(sheet_name)
        writer.book.remove(writer.book.worksheets[idx])
    df.to_excel(writer, sheet_name=sheet_name, index=False)

# Create an Outlook application object and access emails
Outlook = win32com.client.Dispatch("Outlook.Application")
namespace = Outlook.GetNamespace("MAPI")
inbox = namespace.GetDefaultFolder(6)  # 6 refers to the inbox
bhCatBondFolder = inbox.Folders["BH Cat Bond"]
rbcCatBondFolder = inbox.Folders["RBC Database"]

emails = []
sorted_emails = []

for message in bhCatBondFolder.Items:
    if message.UnRead:
        subject = message.Subject
        body = message.Body
        received_time = message.ReceivedTime
        formatted_time = received_time.strftime('%Y-%m-%d %H:%M:%S')
        extracted_body = extract_content_up_to_marker(body, "Craig Bonder")

        emails.append({
            "Timestamp": formatted_time,
            "Subject": subject,
            "Content": extracted_body
        })

        lines = extracted_body.split('\n')
        for line in lines:
            entries = parse_line(line.strip())
            sorted_emails.extend(entries)

        # message.UnRead = False  # Uncomment to mark as read
        # message.Save()  # Uncomment to save the state

emails_df = pd.DataFrame(emails)
sorted_df = pd.DataFrame(sorted_emails)

# File path for the Excel file
excel_file_path = "//ad-its.credit-agricole.fr/Amundi_Boston/Homedirs/buonomo/@Config/Desktop/Outlook Scanner/OrganizedBondEntries.xlsx"

# Open the workbook and remove sheets if they exist
workbook = openpyxl.load_workbook(excel_file_path)
remove_sheet_if_exists(workbook, 'Unread Emails')
remove_sheet_if_exists(workbook, 'Sorted')
workbook.save(excel_file_path)
workbook.close()

# Save to Excel with two sheets
with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
    emails_df.to_excel(writer, sheet_name='Unread Emails', index=False)
    sorted_df.to_excel(writer, sheet_name='Sorted', index=False)


Please adjust the pattern formats and add new pattern formats to account for all of the possibilities before which aren't currently being correctly processed.




ERRORS:

Alamo 2023-1 A (011395AJ9) bid at 102.50
Sorting incorrectly
CUSIP is being put in the Name column and Name is being put in the CUSIP column.

name_cusip_bid_at_price



Applies To All: Should not be storing bid/offer for anything.
If there is bid/offer than it must create two lines (dual-action).

size_name_cusip_offered_at_price
Is storing correctly. However it must denote if it is bid or offer rather than storing as bid/offer.


Tailwind 2022-1 B (87403TAE6) bid at 99
Had no match. We need to figure out why.

5mm Tailwind 2022-1 C (87403TAF3) 99.20 bid / 99.50 offer
size_name_cusip_bid_offer
Stores everything correctly except it is only stored as one line.
Additionally the price for the bid is not being stored.
No second line was created with action bid and bid price.
Actual Output:
Tailwind 2022-1 C	| 5mm	| 87403TAF3	| bid/offer	| 99.50

Expected Output:
Tailwind 2022-1 C	| 5mm	| 87403TAF3	| bid	| 99.20
Tailwind 2022-1 C	| 5mm	| 87403TAF3	| offer	| 99.50


size_name_cusip_bid_offer



size_name_cusip_offered_at_price





Res Re 2020-2 CL3 76120AAB8 bid @ 99.00
Is running with cusip_first_bid_at_price
And Re 2020-2 CL3 76120AAB8
76120AAB8 (the CUSIP) is being attached to the name. Additionally, part of the name is being put within the CUSIP column
Actual Output:
Re 2020-2 CL3 76120AAB8 |		| Res	| bid/offer	| 99.00



5mm Kilimanjaro 2021-1 C (49407PAG5) offered @ 100.10
Ran with name_cusip_offered_at_price_no_size even though it has a size


5mm Kilimanjaro 2021-1 C (49407PAG5) - 99.10 bid / 100.10 offer
Had no match.
