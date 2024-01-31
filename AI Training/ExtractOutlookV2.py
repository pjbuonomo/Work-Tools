import win32com.client
import pandas as pd
import re
from datetime import datetime
import openpyxl

def extract_content_up_to_marker(body, marker):
    marker_position = body.find(marker)
    if marker_position != -1:
        return body[:marker_position].strip()
    else:
        return body.strip()

def remove_sheet_if_exists(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

import re

def parse_line(line):
    patterns = {
        "size_name_cusip_offered_at_price": r"(\d+(\.\d+)?(mm|m|k))\s+([\w\s-]+?)\s+\((\w+)\)\s+offered\s+(?:@|at)\s+(\d*\.\d+)",
        "name_cusip_bid_at_price": r"([\w\s-]+?)\s+\((\w+)\)\s+bid\s+(?:@|at)\s+(\d+\.\d+)",
        "size_name_cusip_bid_offer": r"(\d+(\.\d+)?(mm|m|k))\s+([\w\s-]+?)\s+\((\w+)\)\s+(\d+\.\d+)\s+bid\s+/\s+(\d+\.\d+)\s+offer",
        "name_cusip_offered_at_price_no_size": r"([\w\s-]+?)\s+\((\w+)\)\s+offered\s+(?:@|at)\s+(\d+\.\d+)",
        "bid_price_for_name_cusip": r"(\d+\.\d+)\s+bid\s+for\s+([\w\s-]+?)\s+\((\w+)\)",
        "cusip_first_bid_at_price": r"(\w+)\s+([\w\s-]+?)\s+bid\s+(?:@|at)\s+(\d+\.\d+)"
    }

    default_dict = {"Name": "", "Size": "", "CUSIP": "", "Actions": "", "Price": "", "Error": line}
    entries = []

    for key, pattern in patterns.items():
        if re.match(pattern, line):
            for match in re.finditer(pattern, line):
                groups = match.groups()
                if key in ["size_name_cusip_offered_at_price", "size_name_cusip_bid_offer"]:
                    entries.append({"Name": groups[3].strip(), "Size": groups[0], "CUSIP": groups[4], "Actions": "bid/offer", "Price": groups[-1], "Error": ""})
                elif key in ["name_cusip_bid_at_price", "name_cusip_offered_at_price_no_size", "cusip_first_bid_at_price"]:
                    entries.append({"Name": groups[1].strip(), "CUSIP": groups[0], "Actions": "bid/offer", "Price": groups[2], "Error": ""})
                elif key == "bid_price_for_name_cusip":
                    entries.append({"Name": groups[1].strip(), "CUSIP": groups[2], "Actions": "bid", "Price": groups[0], "Error": ""})

    return entries if entries else [default_dict]

# Continue with your existing code...



def write_df_to_excel(writer, df, sheet_name):
    if sheet_name in writer.book.sheetnames:
        idx = writer.book.sheetnames.index(sheet_name)
        writer.book.remove(writer.book.worksheets[idx])
    df.to_excel(writer, sheet_name=sheet_name, index=False)

# Create an Outlook application object and access emails
Outlook = win32com.client.Dispatch("Outlook.Application")
namespace = Outlook.GetNamespace("MAPI")
inbox = namespace.GetDefaultFolder(6)  # 6 refers to the inbox
bhCatBondFolder = inbox.Folders["BH Cat Bond"]
rbcCatBondFolder = inbox.Folders["RBC Database"]

emails = []
sorted_emails = []

for message in bhCatBondFolder.Items:
    if message.UnRead:
        subject = message.Subject
        body = message.Body
        received_time = message.ReceivedTime
        formatted_time = received_time.strftime('%Y-%m-%d %H:%M:%S')
        extracted_body = extract_content_up_to_marker(body, "Craig Bonder")

        emails.append({
            "Timestamp": formatted_time,
            "Subject": subject,
            "Content": extracted_body
        })

        lines = extracted_body.split('\n')
        for line in lines:
            entries = parse_line(line.strip())
            sorted_emails.extend(entries)

        # message.UnRead = False  # Uncomment to mark as read
        # message.Save()  # Uncomment to save the state

emails_df = pd.DataFrame(emails)
sorted_df = pd.DataFrame(sorted_emails)

# File path for the Excel file
excel_file_path = "//ad-its.credit-agricole.fr/Amundi_Boston/Homedirs/buonomo/@Config/Desktop/Outlook Scanner/OrganizedBondEntries.xlsx"

# Open the workbook and remove sheets if they exist
workbook = openpyxl.load_workbook(excel_file_path)
remove_sheet_if_exists(workbook, 'Unread Emails')
remove_sheet_if_exists(workbook, 'Sorted')
workbook.save(excel_file_path)
workbook.close()

# Save to Excel with two sheets
with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
    emails_df.to_excel(writer, sheet_name='Unread Emails', index=False)
    sorted_df.to_excel(writer, sheet_name='Sorted', index=False)


Please adjust the pattern formats and add new pattern formats to account for all of the possibilities before which aren't currently being correctly processed.


2.25mm Gateway 2023-3 A (36779CAF3) offered @ 107.10
2.25mm Gateway 2023-3 A (36779CAF3) offered @ 107.10
Tailwind 2022-1 B (87403TAE6) bid at 99
2.25mm Gateway 2023-3 A (36779CAF3) offered @ 107.00
Tailwind 2022-1 B (87403TAE6) bid at 99
2.25mm Gateway 2023-3 A (36779CAF3) 102 bid / 107.00 offer
74.50 bid for 3264 Re 2022-1 (88577CAB7)
64.50 bid for Herbie 2021-1 A (42703VAE3)
3mm Mystic 2021-2 B (62865LAC1) - 97.35 bid / 98.10 offer
1.25mm Gateway 2023-1 A (36779CAC0) offered @ 109.90
1mm Purple Re 2023-1 A (74639NAA1) - 102.35 bid / 103.10 offer
74.50 bid for 3264 Re 2022-1 (88577CAB7)
64.50 bid for Herbie 2021-1 A (42703VAE3)
2mm Northshore 2022-1 A (666842AE9) - 103.50 bid / 104 offer
2.5mm Galileo 2023-1 A (36354TAN2)- 101.25 bid / 101.60 offer
2.5mm Galileo 2023-1 B (36354TAP7) -101.25 bid / 101.60 offer
1mm Purple Re 2023-1 A (74639NAA1) - 102.35 bid / 103.10 offer
74.50 bid for 3264 Re 2022-1 (88577CAB7)
64.50 bid for Herbie 2021-1 A (42703VAE3)
2mm Northshore 2022-1 A (666842AE9) - 103.50 bid / 104 offer
1.75mm Galileo 2023-1 B (36354TAP7) offered @ 101.50
2.5mm Ursa 2023-1 AA (90323WAP5) offered @ 100.60
2.5mm Ursa 2023-1 D (90323WAQ3) offered @ 100.85
74.50 bid for 3264 Re 2022-1 (88577CAB7)
64.50 bid for Herbie 2021-1 A (42703VAE3)
1mm Purple Re 2023-1 A (74639NAA1) - 102.35 bid / 103.10 offer
500k Hypatia Re 2023-1 A (44914CAC0)- 104.50 bid / 106.20 offer
500k Hypatia Re 2023-1 A (44914CAC0)- 104.50 bid / 106.20 offer
49407PAJ9 Kilimanjaro Re III 2021-2 B-2 bid @ 96.5
613752AB0 Montoya Re 2022-2 A bid @ 109.5
869255AA7 Sussex Capital 2021-1 bid @ 99.4
87403TAD8 Tailwind Re 2021-1 A bid @ 97.65
888329AB5 Titania Re 2021-2 A bid @ 97.8
888329AC3 Titania Re 2023-1 A bid @ 108.6
Res Re 2020-2 CL4 76120AAC6 bid @ 98.50
Res Re 2021-2 CL3 76114NAF9 bid @ 96.75
Res Re 2021-1 CL12 76120AAB8 bid @ 99.00
Res Re 2020-2 CL3 76120AAB8 bid @ 99.00
Res Re 2020-2 CL4 76120AAC6 bid @ 98.50
Res Re 2021-2 CL3 76114NAF9 bid @ 96.75
425k Mayflower 2023-1 (57839MAA6) - 101.70 bid / 102.35 offer
425k Mayflower 2023-1 (57839MAA6) - 101.90 bid / 102.35 offer
5mm Kilimanjaro 2021-1 C (49407PAG5) - 99.10 bid / 100.10 offer
74.50 bid for 3264 Re 2022-1 (88577CAB7)
64.50 bid for Herbie 2021-1 A (42703VAE3)
2mm Mystic Re IV 2021-2 A (62865LAB3) - 98.25 bid / 99.10 offer
3.25mm Res Re 2023 II 5 (76090WAC4) offered @ 100.85
3.75mm Titania Re 2023-1 A (888329AC3) offered @ 108.65
6.5mm Tailwind 2022-1 C (87403TAF3) offered @ 100.10
4mm Vista 2022-1 A (92840DAB8) - 104.10 bid / 104.50 offer
3mm Hypatia 2023-1 A (44914CAC0)- 105.00 bid / 105.90 offer
Alamo 2023-1 A (011395AJ9) bid at 102.50
Blue Sky 2023-1 (XS2728630596) bid at 100.15
Bonanza 2022-1 A (09785EAJ0) bid at 90.00
Bonanza 2023-1 A (09785EAK7) bid at 99.90
Citrus 2023-1 B (177510AM6) bid at 102.40
Easton 2024-1 A (27777AAA9) bid at 100.25
First Coast 2021-1 (31971CAA1) bid at 96.15
First Coast 2023-1 (31969UAA5) bid at 101.10
Galileo 2023-1 B (36354TAP7) bid at 100.25
Galileo 2023-1 A (36354TAN2) bid at 100.25
Hypatia 2023-1 A (44914CAC0) bid at 104.35
Hexagon 2023-1 A (428270AA0) bid at 100.50
Lightning 2023-1 A (532242AA2) bid at 106.30
Matterhorn 2022-I B (577092AQ2) bid at 98.50
Merna 2022-2A (59013MAF9) bid at 98.65
Merna 2023-2 A (59013MAJ1) bid at 104.35
Mona Lisa 2023-1 B (608800AG3) bid at 107.75
Montoya 2022-2 (613752AB0) bid at 108.60
Montoya 2024-1 A (613752AC8) bid at 101.00
Ocelot 2023-1 A (675951AA5) bid at 100.30
Residential Re 2023-2 5 (76090WAC4) bid at 100.45
Tailwind 2022-1 B (87403TAE6) bid at 99
Tailwind 2022-1 C (87403TAE) bid at 99.20
Titania 2021-1 A (888329AA7) bid at 100.40
Titania 2021-2 A (888329AB5) bid at 97.50
Titania 2023-1 A (888329AC3) bid at 108.50
Ursa 2023-1 C (90323WAM2) bid at 100.45
Ursa 2023-3 D (90323WAQ3) bid at 100.35
5mm Tailwind 2022-1 C (87403TAF3) 99.20 bid / 99.50 offer
500k Res Re 2020-I 13 (76124AAB4) offered @ 98.30
5mm Merna 2022-1 (59013MAE2) offered @ 100.50
5mm Tailwind 2022-1 C (87403TAF3) **BH trades**
500k Riverfront 2021-1 A (76870YAD4) 97.90 bid / 98.60 offer
2.25mm Gateway 2023-3 A (36779CAF3) offered @ 107.00
Alamo 2023-1 A (011395AJ9) bid at 102.50
Blue Sky 2023-1 (XS2728630596) bid at 100.60
Bonanza 2022-1 A (09785EAJ0) bid at 90.00
Bonanza 2023-1 A (09785EAK7) bid at 99.90
Citrus 2023-1 B (177510AM6) bid at 102.40
Easton 2024-1 A (27777AAA9) bid at 100.25
First Coast 2021-1 (31971CAA1) bid at 96.15
First Coast 2023-1 (31969UAA5) bid at 101.10
Galileo 2023-1 B (36354TAP7) bid at 100.25
Galileo 2023-1 A (36354TAN2) bid at 100.25
Hypatia 2023-1 A (44914CAC0) bid at 104.35
Hexagon 2023-1 A (428270AA0) bid at 100.90
Lightning 2023-1 A (532242AA2) bid at 106.30
Matterhorn 2022-I B (577092AQ2) bid at 98.50
Merna 2022-2A (59013MAF9) bid at 98.65
Merna 2023-2 A (59013MAJ1) bid at 104.35
Mona Lisa 2023-1 B (608800AG3) bid at 107.75
Montoya 2022-2 (613752AB0) bid at 108.60
Montoya 2024-1 A (613752AC8) bid at 101.00
Ocelot 2023-1 A (675951AA5) bid at 100.30
Residential Re 2023-2 5 (76090WAC4) bid at 100.45
Tailwind 2022-1 B (87403TAE6) bid at 99
Tailwind 2022-1 C (87403TAE) bid at 99.20
Titania 2021-1 A (888329AA7) bid at 100.40
Titania 2021-2 A (888329AB5) bid at 97.50
Titania 2023-1 A (888329AC3) bid at 108.50
Ursa 2023-1 C (90323WAM2) bid at 100.45
Ursa 2023-3 D (90323WAQ3) bid at 100.35
5mm Merna 2022-1 (59013MAE2) offered @ 100.50
500k Riverfront 2021-1 A (76870YAD4) 97.90 bid / 98.60 offer
2.25mm Gateway 2023-3 A (36779CAF3) 102 bid / 107.00 offer
3mm Mystic 2021-2 B (62865LAC1) offered @ 98.10
3mm Gateway 2022-1 A (36779CAA4) offered @ 102.00
1m Purple Re 2023-1 A (74639NAA1) offered at 103.10
