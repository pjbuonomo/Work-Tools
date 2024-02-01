import win32com.client
import pandas as pd
import re
from datetime import datetime
import openpyxl

def extract_content_up_to_marker(body, marker):
    marker_position = body.find(marker)
    if marker_position != -1:
        return body[:marker_position].strip()
    else:
        return body.strip()

def remove_sheet_if_exists(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

def parse_line(line):
    patterns = {
        "size_name_cusip_offered_at_price": r"(\d+(\.\d+)?(mm|m|k))\s+([\w\s-]+?)\s+\((\w+)\)\s+offered\s+(?:@|at)\s+(\d*\.\d+)",
        "name_cusip_bid_at_price": r"([\w\s-]+?)\s+\((\w+)\)\s+bid\s+(?:@|at)\s+(\d+\.\d+)",
        "size_name_cusip_bid_offer": r"(\d+(\.\d+)?(mm|m|k))\s+([\w\s-]+?)\s+\((\w+)\)\s+(\d+\.\d+)\s+bid\s+/\s+(\d+\.\d+)\s+offer",
        "name_cusip_offered_at_price_no_size": r"([\w\s-]+?)\s+\((\w+)\)\s+offered\s+(?:@|at)\s+(\d+\.\d+)",
        "bid_price_for_name_cusip": r"(\d+\.\d+)\s+bid\s+for\s+([\w\s-]+?)\s+\((\w+)\)",
        "cusip_first_bid_at_price": r"(\w+)\s+([\w\s-]+?)\s+bid\s+(?:@|at)\s+(\d+\.\d+)"
    }

default_dict = {"Name": "", "Size": "", "CUSIP": "", "Actions": "", "Price": "", "Sentence": line, "Function": "No Match", "Error": line}
entries = []

for key, pattern in patterns.items():
    if re.search(pattern, line):
        for match in re.finditer(pattern, line):
            groups = match.groups()
            entry = {"Sentence": line, "Function": key, "Error": ""}

            if key == "size_name_cusip_offered_at_price":
                action = "offer"
                entry.update({"Name": groups[3].strip(), "Size": groups[0], "CUSIP": groups[4], "Actions": action, "Price": groups[5]})

            elif key == "name_cusip_bid_at_price":
                action = "bid"
                entry.update({"Name": groups[0].strip(), "CUSIP": groups[1], "Actions": action, "Price": groups[2]})

            elif key == "size_name_cusip_bid_offer":
                entry_bid = entry.copy()
                entry_bid.update({"Name": groups[3].strip(), "Size": groups[0], "CUSIP": groups[4], "Actions": "bid", "Price": groups[5]})
                entries.append(entry_bid)

                entry_offer = entry.copy()
                entry_offer.update({"Name": groups[3].strip(), "Size": groups[0], "CUSIP": groups[4], "Actions": "offer", "Price": groups[5]})
                entries.append(entry_offer)
                continue

            elif key == "name_cusip_offered_at_price_no_size":
                action = "offer"
                entry.update({"Name": groups[0].strip(), "CUSIP": groups[1], "Actions": action, "Price": groups[2]})

            elif key == "bid_price_for_name_cusip":
                entry.update({"Name": groups[1].strip(), "CUSIP": groups[2], "Actions": "bid", "Price": groups[0]})

            elif key == "cusip_first_bid_at_price":
                action = "bid"
                entry.update({"Name": groups[1].strip(), "CUSIP": groups[0], "Actions": action, "Price": groups[2]})

            entries.append(entry)

return entries if entries else [default_dict]



def write_df_to_excel(writer, df, sheet_name):
    if sheet_name in writer.book.sheetnames:
        idx = writer.book.sheetnames.index(sheet_name)
        writer.book.remove(writer.book.worksheets[idx])
    df.to_excel(writer, sheet_name=sheet_name, index=False)

# Create an Outlook application object and access emails
Outlook = win32com.client.Dispatch("Outlook.Application")
namespace = Outlook.GetNamespace("MAPI")
inbox = namespace.GetDefaultFolder(6)  # 6 refers to the inbox
bhCatBondFolder = inbox.Folders["BH Cat Bond"]
rbcCatBondFolder = inbox.Folders["RBC Database"]

emails = []
sorted_emails = []

for message in bhCatBondFolder.Items:
    if message.UnRead:
        subject = message.Subject
        body = message.Body
        received_time = message.ReceivedTime
        formatted_time = received_time.strftime('%Y-%m-%d %H:%M:%S')
        extracted_body = extract_content_up_to_marker(body, "Craig Bonder")

        emails.append({
            "Timestamp": formatted_time,
            "Subject": subject,
            "Content": extracted_body
        })

        lines = extracted_body.split('\n')
        for line in lines:
            entries = parse_line(line.strip())
            sorted_emails.extend(entries)

        # message.UnRead = False  # Uncomment to mark as read
        # message.Save()  # Uncomment to save the state

emails_df = pd.DataFrame(emails)
sorted_df = pd.DataFrame(sorted_emails)

# File path for the Excel file
excel_file_path = "//ad-its.credit-agricole.fr/Amundi_Boston/Homedirs/buonomo/@Config/Desktop/Outlook Scanner/OrganizedBondEntries.xlsx"

# Open the workbook and remove sheets if they exist
workbook = openpyxl.load_workbook(excel_file_path)
remove_sheet_if_exists(workbook, 'Unread Emails')
remove_sheet_if_exists(workbook, 'SortedV3')
workbook.save(excel_file_path)
workbook.close()

# Save to Excel with two sheets
with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
    emails_df.to_excel(writer, sheet_name='Unread Emails', index=False)
    sorted_df.to_excel(writer, sheet_name='SortedV3', index=False)
